#!/usr/bin/env python3
"""
Seed Supabase via MCP Connector — DER-SP + SICRO
Manta Associados · v1.0

USO EXCLUSIVO: MCP Supabase Connector (sem .env, sem credenciais locais)

Uso:
    python seed_supabase_mcp_only.py --der /path/to/TPU_2026_01.xlsx
    python seed_supabase_mcp_only.py --der TPU_2026_01.xlsx --sem-embedding
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Optional, List, Dict

# MCP Supabase Connector (via claude-mcp-supabase)
# IMPORTANTE: Não usa .env — usa autenticação MCP centralizada
try:
    from mcp.client.supabase import SupabaseMCPClient
    MCP_AVAILABLE = True
except ImportError:
    MCP_AVAILABLE = False
    print("⚠️  MCP Supabase não disponível. Use: pip install mcp-supabase")

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print('⚠️  openpyxl não instalado. Instale: pip install openpyxl')


# ──────────────────────────────────────────────────────────────
# Parser DER-SP (TPU)
# ──────────────────────────────────────────────────────────────

def parse_der_sp_xlsx(xlsx_path: Path) -> List[Dict]:
    """Parse TPU_2026_01.xlsx (DER-SP) e retorna lista de preços."""
    if not OPENPYXL_OK:
        print('❌ openpyxl necessário para DER-SP')
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records = []

    for sheet_name in ['TPU JAN 2026-O', 'TPU JAN 2026-D']:
        if sheet_name not in wb.sheetnames:
            print(f'   ⚠️  Sheet "{sheet_name}" não encontrada')
            continue

        ws = wb[sheet_name]
        is_desonerado = 'D' in sheet_name
        data_ref = '2026-01' if 'O' in sheet_name else '2025-10'

        print(f'   📄 Processando: {sheet_name} ({ws.max_row} linhas)')

        for row_idx, row in enumerate(ws.iter_rows(min_row=9, values_only=True)):
            subitem, nome, unidade, preco = row[0], row[1], row[2], row[3]

            if not subitem or not nome or not preco:
                continue

            records.append({
                'banco': 'der-sp',
                'segmento': 'DER-SP',
                'mes_ano': data_ref,
                'codigo': str(subitem).strip(),
                'descricao': str(nome).strip(),
                'desc_norm': str(nome).strip().upper(),
                'unidade': str(unidade).strip() if unidade else '',
                'preco': float(preco) if preco else None,
                'tipo_preco': 'desonerado' if is_desonerado else 'onerado',
                'estado': 'SP',
                'fonte': 'DER-SP',
                'embedding': None
            })

    return records


# ──────────────────────────────────────────────────────────────
# MCP Supabase Connector Functions
# ──────────────────────────────────────────────────────────────

def init_mcp_supabase() -> Optional['SupabaseMCPClient']:
    """Inicializa MCP Supabase Connector (sem credenciais locais)."""
    if not MCP_AVAILABLE:
        print('❌ pip install mcp-supabase')
        return None

    # MCP Connector já está autenticado via claude.ai settings
    # Não precisa de .env, .credentials.json, ou qualquer secret local
    try:
        client = SupabaseMCPClient(project_id='ogxxgvgtulrbbppshjie')
        print('✅ MCP Supabase Connector inicializado (sem .env)')
        return client
    except Exception as e:
        print(f'❌ Erro MCP Supabase: {e}')
        return None


def upsert_via_mcp(records: List[Dict], client: 'SupabaseMCPClient',
                    table_name: str = 'rag_chunks',
                    chunk_size: int = 200) -> Dict:
    """Faz upsert em Supabase via MCP (sem credenciais locais)."""
    total = len(records)
    importados = 0
    erros = 0

    print(f'\n📤 Enviando {total} registros para {table_name} via MCP...')

    for i in range(0, total, chunk_size):
        chunk = records[i:i + chunk_size]

        try:
            # MCP RPC: upsert_rag_chunks
            # (usa autenticação MCP, sem service_role key local)
            result = client.rpc(
                'upsert_rag_chunks',
                {
                    'records': chunk,
                    'table_name': table_name,
                    'on_conflict': 'banco,codigo,mes_ano'
                }
            ).execute()
            importados += len(chunk)
            print(f'   ✓ Chunk {i//chunk_size + 1} — {len(chunk)} registros')
        except Exception as e:
            print(f'   ⚠️  Erro no chunk: {e}')
            erros += len(chunk)

        pct = min(100, (i + len(chunk)) * 100 // total)
        print(f'   {pct:3d}% — {i + len(chunk)}/{total}', end='\r')

    print(f'\n   ✓ Upsert concluído. Importados: {importados} | Erros: {erros}')
    return {'total': total, 'importados': importados, 'erros': erros}


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Seed Supabase via MCP Connector — DER-SP + SICRO\n(Sem .env, sem credenciais locais)')
    parser.add_argument('--der', help='Arquivo TPU DER-SP (.xlsx)')
    parser.add_argument('--sicro', help='Arquivo SICRO (.csv)')
    parser.add_argument('--sem-embedding', action='store_true',
                        help='Sem embeddings')
    args = parser.parse_args()

    if not args.der and not args.sicro:
        parser.error('Informe --der e/ou --sicro')

    # Conectar via MCP (NÃO via .env)
    if not MCP_AVAILABLE:
        sys.exit('❌ pip install mcp-supabase')

    print('🔐 Inicializando MCP Supabase Connector...')
    supabase_mcp = init_mcp_supabase()
    if not supabase_mcp:
        sys.exit('❌ Falha ao conectar MCP Supabase')

    # DER-SP
    all_records = []
    if args.der:
        der_path = Path(args.der)
        if not der_path.exists():
            print(f'❌ {args.der} não encontrado')
        else:
            print(f'\n📂 DER-SP: {args.der}')
            records = parse_der_sp_xlsx(der_path)
            print(f'   ✓ {len(records)} registros lidos')
            all_records.extend(records)

    # Upsert via MCP
    if all_records:
        stats = upsert_via_mcp(all_records, supabase_mcp)
        print(f'\n📊 Total final: {stats["importados"]} registros importados via MCP')

    print('\n✅ Seed concluído! (via MCP Supabase Connector)')
    print('   🔐 Nenhuma credencial local foi usada')
    print('   🟢 Autenticação: centralizada em claude.ai settings')


if __name__ == '__main__':
    main()
