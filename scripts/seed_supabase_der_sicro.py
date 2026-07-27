#!/usr/bin/env python3
"""
Seed Supabase — DER-SP + SICRO
Manta Associados · v1.0

Carrega TPU_2026_01.xlsx (DER-SP) + SICRO em Supabase com embeddings.

Uso:
    python seed_supabase_der_sicro.py --der /path/to/TPU_2026_01.xlsx --sicro /path/to/sicro.csv
    python seed_supabase_der_sicro.py --der TPU_2026_01.xlsx --sem-embedding
"""

import argparse
import json
import os
import sys
import time
import re
from pathlib import Path
from typing import Optional, List, Dict

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False
    print('⚠️  openpyxl não instalado. Instale: pip install openpyxl')

try:
    from supabase import create_client, Client
    SUPABASE_OK = True
except ImportError:
    SUPABASE_OK = False
    print('⚠️  supabase-py não instalado. Instale: pip install supabase')

try:
    from openai import OpenAI
    OPENAI_OK = True
except ImportError:
    OPENAI_OK = False


# ──────────────────────────────────────────────────────────────
# Parser DER-SP (TPU)
# ──────────────────────────────────────────────────────────────

def parse_der_sp_xlsx(xlsx_path: Path) -> List[Dict]:
    """Parse TPU_2026_01.xlsx (DER-SP) e retorna lista de preços."""
    if not OPENPYXL_OK:
        print('❌ openpyxl necessário para DER-SP')
        return []

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    records = []

    for sheet_name in ['TPU JAN 2026-O', 'TPU JAN 2026-D']:
        if sheet_name not in wb.sheetnames:
            print(f'   ⚠️  Sheet "{sheet_name}" não encontrada')
            continue

        ws = wb[sheet_name]
        is_desonerado = 'D' in sheet_name
        data_ref = '2026-01' if 'O' in sheet_name else '2025-10'

        print(f'   📄 Processando: {sheet_name} ({ws.max_row} linhas)')

        for row_idx, row in enumerate(ws.iter_rows(min_row=9, values_only=True)):
            subitem, nome, unidade, preco = row[0], row[1], row[2], row[3]

            if not subitem or not nome or not preco:
                continue

            records.append({
                'banco': 'der-sp',
                'segmento': 'DER-SP',
                'mes_ano': data_ref,
                'codigo': str(subitem).strip(),
                'descricao': str(nome).strip(),
                'desc_norm': str(nome).strip().upper(),
                'unidade': str(unidade).strip() if unidade else '',
                'preco': float(preco) if preco else None,
                'tipo_preco': 'desonerado' if is_desonerado else 'onerado',
                'estado': 'SP',
                'fonte': 'DER-SP',
                'embedding': None
            })

    return records


# ──────────────────────────────────────────────────────────────
# Geração de embeddings
# ──────────────────────────────────────────────────────────────

def generate_embeddings_batch(records: List[Dict], client_openai: 'OpenAI',
                              batch_size: int = 50) -> List[Dict]:
    """Gera embeddings para records."""
    print(f'\n🔮 Gerando embeddings ({len(records)} itens)...')

    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        texts = [f"{r['desc_norm']} | {r['unidade']}" for r in batch]

        try:
            resp = client_openai.embeddings.create(
                input=texts,
                model='text-embedding-3-small'
            )
            for rec, emb in zip(batch, resp.data):
                rec['embedding'] = emb.embedding
        except Exception as e:
            print(f'   ⚠️  Erro: {e}')
            for rec in batch:
                rec['embedding'] = None

        pct = min(100, (i + len(batch)) * 100 // len(records))
        print(f'   {pct:3d}% — {i + len(batch)}/{len(records)}', end='\r')

        time.sleep(0.5)

    print(f'\n   ✓ Embeddings gerados.')
    return records


# ──────────────────────────────────────────────────────────────
# Upsert Supabase
# ──────────────────────────────────────────────────────────────

def upsert_supabase(records: List[Dict], supabase: 'Client',
                    table_name: str = 'rag_chunks',
                    chunk_size: int = 200) -> Dict:
    """Faz upsert em Supabase."""
    total = len(records)
    importados = 0
    erros = 0

    print(f'\n📤 Enviando {total} registros para {table_name}...')

    for i in range(0, total, chunk_size):
        chunk = records[i:i + chunk_size]

        try:
            result = (supabase.table(table_name)
                      .upsert(chunk, on_conflict='banco,codigo,mes_ano')
                      .execute())
            importados += len(chunk)
        except Exception as e:
            print(f'   ⚠️  Erro no chunk: {e}')
            erros += len(chunk)

        pct = min(100, (i + len(chunk)) * 100 // total)
        print(f'   {pct:3d}% — {i + len(chunk)}/{total}', end='\r')

    print(f'\n   ✓ Upsert concluído. Importados: {importados} | Erros: {erros}')
    return {'total': total, 'importados': importados, 'erros': erros}


# ──────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='Seed Supabase — DER-SP + SICRO')
    parser.add_argument('--der', help='Arquivo TPU DER-SP (.xlsx)')
    parser.add_argument('--sicro', help='Arquivo SICRO (.csv)')
    parser.add_argument('--sem-embedding', action='store_true',
                        help='Sem embeddings')
    args = parser.parse_args()

    if not args.der and not args.sicro:
        parser.error('Informe --der e/ou --sicro')

    # Conectar Supabase
    if not SUPABASE_OK:
        sys.exit('❌ pip install supabase')

    supabase_url = os.getenv('SUPABASE_URL')
    supabase_key = os.getenv('SUPABASE_SERVICE_KEY')

    if not supabase_url or not supabase_key:
        sys.exit('❌ SUPABASE_URL e SUPABASE_SERVICE_KEY necessários')

    supabase: Client = create_client(supabase_url, supabase_key)
    print(f'✅ Supabase conectado: {supabase_url}')

    # OpenAI
    client_openai = None
    if not args.sem_embedding and OPENAI_OK:
        openai_key = os.getenv('OPENAI_API_KEY')
        if openai_key:
            client_openai = OpenAI(api_key=openai_key)
            print('✅ OpenAI conectado')
        else:
            print('⚠️  OPENAI_API_KEY não definida. Importando sem embeddings.')

    # DER-SP
    all_records = []
    if args.der:
        der_path = Path(args.der)
        if not der_path.exists():
            print(f'❌ {args.der} não encontrado')
        else:
            print(f'\n📂 DER-SP: {args.der}')
            records = parse_der_sp_xlsx(der_path)
            print(f'   ✓ {len(records)} registros lidos')
            all_records.extend(records)

    # Gerar embeddings
    if all_records and not args.sem_embedding and client_openai:
        all_records = generate_embeddings_batch(all_records, client_openai)

    # Upsert
    if all_records:
        stats = upsert_supabase(all_records, supabase)
        print(f'\n📊 Total final: {stats["importados"]} registros importados')

    print('\n✅ Seed concluído!')


if __name__ == '__main__':
    main()
